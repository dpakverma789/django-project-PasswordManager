from .password_ecryptor import *
from django.contrib.auth.hashers import check_password
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pytz, random
from .models import Credentials

def home(request):
    if not request.user.is_authenticated:
        return redirect('signin-page')

    data = {'title': 'PassWord Manager', 'header': 'DASHLINE', 'user': request.user}

    if request.POST:
        site = request.POST.get('site')
        username = request.POST.get('text')
        password = request.POST.get('password')

        if not Credentials.objects.filter(website=site, login_user=request.user).exists():
            cred = Credentials(
                website=site,
                username=username,
                login_user=request.user
            )

            while True:
                key = random.randint(100, 999)
                encrypted_text = text_encryption(plain_text=password, salt=key)
                try:
                    text_decryption(encrypted_text_received=encrypted_text)
                except Exception:
                    continue
                else:
                    cred.password = encrypted_text
                    break

            cred.save()
            data.update({'color': '#47d147', 'msg': 'Credentials Saved!'})
        else:
            data.update({'color': '#ffa31a', 'msg': 'Already Exist!'})

    return render(request, 'password_manager.html', {'data': data})

def api(request):
    return render(request, 'api.html')

def recovery(request, website=None):
    if not request.user.is_authenticated:
        return redirect('signin-page')

    header = 'RECOVERY' if 'delete' not in request.path and 'recover' in request.path else 'DELETE'
    data = {
        'title': f'PassWord {header.capitalize()}',
        'header': header,
        'user': request.user,
        'website': website
    }

    if request.method == 'POST':
        if check_password(request.POST.get('pass'), request.user.password):
            site = request.POST.get('site')
            try:
                cred = Credentials.objects.get(website=site, login_user=request.user)

                if header == 'RECOVERY':
                    password_decrypted = text_decryption(encrypted_text_received=cred.password)
                    data = {
                        'website': cred.website,
                        'username': cred.username,
                        'password': password_decrypted,
                        'title': 'PassWord Manager',
                        'header': 'DASHLINE',
                        'user': request.user
                    }
                    return render(request, 'password_manager.html', {'data': data})
                else:
                    cred.delete()
                    data.update({'color': '#9933ff', 'msg': 'Credentials Removed!'})

            except Credentials.DoesNotExist:
                data.update({'msg': 'Website not Found!', 'color': '#ff3333'})
        else:
            messages.error(request, 'Your Password is Incorrect!')

    return render(request, 'recover.html', {'data': data})

def update(request, website=None):
    if not request.user.is_authenticated:
        return redirect('signin-page')

    data = {
        'title': 'PassWord Update',
        'header': 'UPDATE',
        'user': request.user,
        'website': website
    }

    if request.method == 'POST':
        if check_password(request.POST.get('pass'), request.user.password):
            site = request.POST.get('site')
            update_field = request.POST.get('update')
            text_value = request.POST.get('text')

            try:
                cred = Credentials.objects.get(website=site, login_user=request.user)
                is_done, msg, color = False, '', ''

                if update_field == 'website':
                    cred.website = text_value
                    is_done = True
                elif update_field == 'username':
                    cred.username = text_value
                    is_done = True
                elif update_field == 'password':
                    while True:
                        key = random.randint(100, 999)
                        encrypted_text = text_encryption(plain_text=text_value, salt=key)
                        try:
                            text_decryption(encrypted_text_received=encrypted_text)
                        except Exception:
                            continue
                        else:
                            cred.password = encrypted_text
                            break
                    is_done = True

                if is_done:
                    cred.save()
                    color = '#47d147'
                    msg = f"{update_field} has been updated!"
                else:
                    color = '#ff3333'
                    msg = 'Please select the option!'

                data.update({'color': color, 'msg': msg})

            except Credentials.DoesNotExist:
                data.update({'msg': 'Website not Found!', 'color': '#ff3333'})
        else:
            messages.error(request, 'Your Password is Incorrect!')

    return render(request, 'update.html', {'data': data})

def export(request):
    if not request.user.is_authenticated:
        return redirect('signin-page')

    rows = Credentials.objects.filter(login_user=request.user).order_by('website')
    if not rows.exists():
        data = {
            'title': 'PassWord Manager',
            'header': 'DASHLINE',
            'user': request.user,
            'color': '#ff3333',
            'msg': 'Nothing to export'
        }
        return render(request, 'password_manager.html', {'data': data})

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    time_zone = pytz.timezone('Asia/Kolkata')
    time_stamp = datetime.now(time_zone).strftime("%d-%B-%Y--%H-%M")
    file_name = f"{request.user}-{time_stamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"

    # Header
    ws.append(["Website", "Username", "Password"])

    # Rows
    for cred in rows:
        ws.append([cred.website, cred.username, text_decryption(cred.password)])
    wb.save(response)
    return response

def file_import(request):
    if not request.user.is_authenticated:
        return redirect('signin-page')

    data = {
        'title': 'PassWord Manager',
        'header': 'DASHLINE',
        'user': request.user
    }

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']

        # File size limit (250 MB)
        if file.size > 250 * 1024 * 1024:
            data.update({'color': '#ff3333', 'msg': 'File too large (max 250MB).'})
            return render(request, 'import.html', {'data': data})

        # File type check
        if not file.name.endswith('.xlsx'):
            data.update({'color': '#ff3333', 'msg': 'Only .xlsx files are supported.'})
            return render(request, 'import.html', {'data': data})

        try:
            wb = load_workbook(file, data_only=True)
            sheet = wb.active

            imported_count = 0
            skipped_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
                if not row:
                    continue  # skip empty row

                # Normalize → replace None with "" and strip
                row = [str(c).strip() if c is not None else "" for c in row[:3]]

                # Skip fully empty rows
                if not any(row):
                    continue

                website, username, plain_password = row

                # Skip incomplete rows
                if not (website and username and plain_password):
                    continue

                # Check if entry already exists for this user
                if Credentials.objects.filter(
                    website=website,
                    username=username,
                    login_user=request.user
                ).exists():
                    skipped_count += 1
                    continue

                # Create new entry
                Credentials.objects.create(
                    website=website[:100],  # truncate to DB size
                    username=username[:100],
                    password=text_encryption(
                        plain_text=plain_password[:20],  # enforce 20-char max
                        salt=random.randint(100, 999)
                    ),
                    login_user=request.user
                )
                imported_count += 1

            if imported_count > 0:
                msg = f'Imported {imported_count} credentials successfully!'
                if skipped_count:
                    msg += f' Skipped {skipped_count} duplicates.'
                data.update({'color': '#47d147', 'msg': msg})
            else:
                data.update({'color': '#ff9933', 'msg': 'No valid rows found in file.'})

        except Exception as e:
            data.update({'color': '#ff3333', 'msg': f'Error reading file: {e}'})

    return render(request, 'import.html', {'data': data})


def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('signin-page')

    data = {'user': request.user}

    credentials = Credentials.objects.filter(
        login_user=request.user
    ).order_by('website')

    cred_container = []
    for cred in credentials:
        try:
            decrypted_password = text_decryption(encrypted_text_received=cred.password)
        except Exception:
            decrypted_password = "[Error Decrypting]"
        cred_container.append([cred.website, cred.username, decrypted_password])

    return render(
        request,
        'dashboard.html',
        {
            'cred_container': cred_container if cred_container else None,
            'data': data,
        }
    )

def page_not_found_view(request, exception):
    return render(request, '404.html')
